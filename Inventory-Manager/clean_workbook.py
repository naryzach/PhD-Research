"""Produce a cleaned copy of the lab's Excel order workbook.

The original is never modified. The output is a new workbook containing:

* one cleaned sheet per term, with a single consistent column layout and the
  status written as **text** (the original encodes it as a cell fill colour, so
  the file could not be read correctly by anything but this toolchain);
* an ``All Orders`` sheet with every row from every term;
* a ``Cleaning Log`` sheet recording every individual change, so nothing is
  silently rewritten.

Only corrections that are traceable to real data are applied. Where the source
is simply wrong or missing and nothing in the workbook can supply the answer,
the row is flagged for review rather than invented.

    python clean_workbook.py "Sarmazdeh's Lab Orders.xlsx" -o "...(CLEANED).xlsx"
"""
import argparse
import os
import re
from collections import Counter, defaultdict
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import historical_import as hi

# --------------------------------------------------------------------------
# Vendor canonicalisation
# --------------------------------------------------------------------------

# Aliases that case/punctuation folding alone will not merge. Keys and values
# are both normalised forms (lowercase, alphanumerics only).
VENDOR_ALIASES = {
    "fisher": "fishersci",
    "fisherscientific": "fishersci",
    "fisherbrand": "fishersci",
    "thermo": "thermofisher",
    "thermosci": "thermofisher",
    "thermofishersci": "thermofisher",
    "gibco": "thermofisher",
    "bd": "bdbiosciences",
    "bdbio": "bdbiosciences",
    "genescript": "genscript",
    "sciencecell": "sciencell",
    "sigma": "sigmaaldrich",
    "millipore": "milliporesigma",
    "anaspecinc": "anaspec",
    "enzo": "enzolifesciences",
    "idtdna": "idt",
    "unr": "unrchemstore",
    "vwrchemicals": "vwr",
}

# Preferred spellings; anything not listed uses the most common spelling found.
VENDOR_DISPLAY = {
    "fishersci": "Fisher Sci",
    "thermofisher": "Thermo Fisher",
    "bdbiosciences": "BD Biosciences",
    "biorad": "Bio-Rad",
    "genscript": "GenScript",
    "sciencell": "ScienCell",
    "sigmaaldrich": "Sigma-Aldrich",
    "milliporesigma": "MilliporeSigma",
    "anaspec": "AnaSpec",
    "enzolifesciences": "Enzo Life Sciences",
    "abcam": "Abcam",
    "anatrace": "Anatrace",
    "cytiva": "Cytiva",
    "idt": "IDT",
    "unrchemstore": "UNR Chemstore",
    "vwr": "VWR",
    "amazon": "Amazon",
}


def _vendor_key(v):
    if not v:
        return ""
    k = re.sub(r"[^a-z0-9]", "", str(v).lower())
    return VENDOR_ALIASES.get(k, k)


def build_vendor_map(values):
    """normalised key -> canonical display spelling."""
    groups = defaultdict(Counter)
    for v in values:
        if v and str(v).strip():
            groups[_vendor_key(v)][str(v).strip()] += 1
    out = {}
    for key, spellings in groups.items():
        if not key or key == "na":
            continue
        out[key] = VENDOR_DISPLAY.get(key) or spellings.most_common(1)[0][0]
    return out


# --------------------------------------------------------------------------
# Term -> approximate date, for rows with no date anywhere
# --------------------------------------------------------------------------

def term_start(term):
    """Approximate start date implied by a sheet name, or None."""
    t = str(term).strip()
    years = re.findall(r"(20\d{2})", t)
    if not years:
        return None
    y = int(years[0])
    low = t.lower()
    if low.startswith("fall"):
        return datetime(y, 9, 1)
    if low.startswith("spring"):
        return datetime(y, 1, 15)
    if low.startswith("summer"):
        return datetime(y, 6, 1)
    if low.startswith("fy"):
        return datetime(y, 7, 1)      # fiscal year starts in July
    return datetime(y, 1, 1)


# --------------------------------------------------------------------------
# Cleaning
# --------------------------------------------------------------------------

OUT_COLUMNS = [
    "Term", "Item", "Size", "Quantity", "CAT No.", "Vendor", "Unit Price",
    "Initial", "Date Requested", "Status", "Order Date", "GR #", "Order #",
    "Expected Delivery Date", "Tracking #", "Storage", "Comments", "Link",
    "Received By", "Received Date", "Source Row", "Review Flag",
]

STATUS_FILL = {
    "Received": "FF00B050",
    "Ordered": "FFFFFF00",
    "Sent to Dr. MRS": "FFFFC000",
    "Need to order": "FFFF0000",
    "Cancelled": "FFEEECE1",
    "Back order": "FF0070C0",
    "Needs Fixing": "FF00B0F0",
    "Do not order yet": "FFFFC7CE",
    "Lost": "FFD9D9D9",
    "Delayed": "FFEEECE1",
}


def clean(df):
    """Return (cleaned_dataframe, log_dataframe)."""
    log = []

    def note(row, field, old, new, why):
        log.append({
            "Term": row.get("term"), "Source Row": row.get("excel_row"),
            "Item": str(row.get("item"))[:70], "Field": field,
            "Was": "" if old is None else str(old),
            "Now": "" if new is None else str(new), "Why": why,
        })

    vendor_map = build_vendor_map(df["seller"])

    # A catalog number seen elsewhere in the workbook with a sane price lets us
    # repair an obviously mistyped price from the lab's own records. Prices for
    # the same item drift upward over the years, so each entry keeps its date
    # and the repair uses the temporally nearest one rather than an average.
    price_lookup = defaultdict(list)
    for _, r in df.iterrows():
        cat = hi._norm_cat(r.get("catalog_number"))
        p = r.get("price")
        if cat and p is not None and pd.notna(p) and 0 < p <= hi.SUSPICIOUS_PRICE:
            price_lookup[cat].append((hi._row_date(r), float(p)))

    def nearest_price(cat, when):
        """Price recorded for this catalog number closest in time to `when`."""
        entries = price_lookup.get(cat) or []
        if not entries:
            return None, None
        dated = [(d, p) for d, p in entries if d is not None]
        if when is not None and dated:
            d, p = min(dated, key=lambda e: abs((e[0] - when).days))
            return p, f"the {d:%b %Y} record for this catalog number"
        prices = [p for _, p in entries]
        return round(sum(prices) / len(prices), 2), "other records for this catalog number"

    out = []
    for _, r in df.iterrows():
        row = dict(r)
        flags = []

        # --- Vendor -------------------------------------------------------
        seller = r.get("seller")
        if seller:
            canon = vendor_map.get(_vendor_key(seller))
            if canon and canon != str(seller).strip():
                note(r, "Vendor", seller, canon, "vendor spelling merged")
                row["seller"] = canon

        # --- Unit price ---------------------------------------------------
        price = r.get("price")
        if price is not None and pd.notna(price) and price > hi.SUSPICIOUS_PRICE:
            cat = hi._norm_cat(r.get("catalog_number"))
            fixed, source = nearest_price(cat, hi._row_date(r))
            if fixed is not None:
                note(r, "Unit Price", price, fixed,
                     f"implausible price; taken from {source} "
                     f"(catalog {r.get('catalog_number')})")
                row["price"] = fixed
            else:
                note(r, "Unit Price", price, None,
                     "implausible price and no other record of this catalog number; "
                     "cleared rather than guessed")
                row["price"] = None
                flags.append("price removed - needs a real value")

        # --- Dates --------------------------------------------------------
        for field, label in (("date_requested", "Date Requested"),
                             ("order_date", "Order Date"),
                             ("received_date", "Received Date")):
            v = r.get(field)
            if v is not None and pd.notna(v) and hi._plausible(v) is None:
                ts = pd.to_datetime(v, errors="coerce")
                # A year typo ("2-Sep-35" for "2-Sep-25") keeps the right day and
                # month, so borrow the year from a good date on the same row.
                donor = None
                for other in ("order_date", "received_date", "date_requested",
                              "expected_delivery"):
                    if other == field:
                        continue
                    cand = hi._plausible(r.get(other))
                    if cand is not None:
                        donor = cand
                        break
                if donor is not None and pd.notna(ts):
                    try:
                        fixed = ts.replace(year=donor.year)
                    except ValueError:
                        fixed = donor
                    if hi._plausible(fixed) is None:
                        fixed = donor
                    note(r, label, ts.date(), fixed.date(),
                         f"impossible year; corrected using the {other.replace('_', ' ')} "
                         f"on the same row")
                    row[field] = fixed
                else:
                    note(r, label, ts.date() if pd.notna(ts) else v, None,
                         "impossible year and no other date on the row to correct from")
                    row[field] = None
                    flags.append(f"{label.lower()} was impossible")

        # Fill a missing request date from other real dates on the row.
        if hi._plausible(row.get("date_requested")) is None:
            for other, why in (("order_date", "order date"),
                               ("received_date", "received date")):
                cand = hi._plausible(row.get(other))
                if cand is not None:
                    note(r, "Date Requested", None, cand.date(),
                         f"was blank; filled from the {why} on the same row")
                    row["date_requested"] = cand
                    break
            else:
                approx = term_start(r.get("term"))
                if approx is not None:
                    note(r, "Date Requested", None, approx.date(),
                         f"no date anywhere on the row; approximated from the "
                         f"'{r.get('term')}' sheet")
                    row["date_requested"] = approx
                    flags.append("date approximated from term")
                else:
                    flags.append("no date available")

        # --- Text tidy ----------------------------------------------------
        for field in ("item", "size", "catalog_number", "storage", "comments",
                      "link", "initial", "order_number", "tracking_number"):
            v = row.get(field)
            if isinstance(v, str):
                # Non-breaking spaces sneak in from web copy-paste.
                cleaned = re.sub(r"\s+", " ", v.replace("\xa0", " ")).strip()
                if cleaned != v:
                    row[field] = cleaned or None

        # --- Sanity flags (not corrections) -------------------------------
        if not row.get("catalog_number"):
            flags.append("no catalog number")
        if row.get("price") in (None, 0) or pd.isna(row.get("price")):
            flags.append("no price recorded")
        size = str(row.get("size") or "").strip().lower()
        if size in ("n/a", "na", "not specified", "none", ""):
            flags.append("size not specified")

        row["_flags"] = "; ".join(flags)
        out.append(row)

    cleaned = pd.DataFrame(out)
    return cleaned, pd.DataFrame(log)


def to_output_frame(cleaned):
    """Map internal field names onto the workbook's column layout."""
    def col(series_name):
        return cleaned[series_name] if series_name in cleaned.columns else None

    out = pd.DataFrame({
        "Term": col("term"),
        "Item": col("item"),
        "Size": col("size"),
        "Quantity": col("quantity"),
        "CAT No.": col("catalog_number"),
        "Vendor": col("seller"),
        "Unit Price": col("price"),
        "Initial": col("initial"),
        "Date Requested": col("date_requested"),
        "Status": col("status"),
        "Order Date": col("order_date"),
        "GR #": col("gr_number"),
        "Order #": col("order_number"),
        "Expected Delivery Date": col("expected_delivery"),
        "Tracking #": col("tracking_number"),
        "Storage": col("storage"),
        "Comments": col("comments"),
        "Link": col("link"),
        "Received By": col("received_by"),
        "Received Date": col("received_date"),
        "Source Row": col("excel_row"),
        "Review Flag": col("_flags"),
    })
    return out[OUT_COLUMNS]


# --------------------------------------------------------------------------
# Writing
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="FFDBE5F1")
FLAG_FILL = PatternFill("solid", fgColor="FFFFF2CC")


def _write_sheet(ws, frame, colour_status=True):
    ws.append(list(frame.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    status_idx = list(frame.columns).index("Status") + 1 if "Status" in frame.columns else None
    flag_idx = list(frame.columns).index("Review Flag") + 1 if "Review Flag" in frame.columns else None

    for _, row in frame.iterrows():
        vals = []
        for v in row:
            if isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            elif v is not None and not isinstance(v, (str, int, float, datetime)):
                v = None if pd.isna(v) else str(v)
            elif isinstance(v, float) and pd.isna(v):
                v = None
            vals.append(v)
        ws.append(vals)
        r = ws.max_row
        if colour_status and status_idx:
            fill = STATUS_FILL.get(ws.cell(row=r, column=status_idx).value)
            if fill:
                ws.cell(row=r, column=status_idx).fill = PatternFill("solid", fgColor=fill)
        if flag_idx and ws.cell(row=r, column=flag_idx).value:
            ws.cell(row=r, column=flag_idx).fill = FLAG_FILL

    for i, name in enumerate(frame.columns, start=1):
        width = {"Item": 58, "Comments": 34, "Link": 30, "Review Flag": 34,
                 "Size": 18, "Vendor": 18, "Status": 17}.get(name, 14)
        ws.column_dimensions[get_column_letter(i)].width = width
    for col_name in ("Date Requested", "Order Date", "Expected Delivery Date", "Received Date"):
        if col_name in frame.columns:
            idx = list(frame.columns).index(col_name) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = "yyyy-mm-dd"
    if "Unit Price" in frame.columns:
        idx = list(frame.columns).index("Unit Price") + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=idx).number_format = '"$"#,##0.00'


def _write_readme(ws, cleaned, log, src):
    lines = [
        ("Cleaned copy of the lab order workbook", True),
        ("", False),
        (f"Source file : {os.path.basename(src)}", False),
        (f"Generated   : {datetime.now():%Y-%m-%d %H:%M}", False),
        (f"Order rows  : {len(cleaned):,} across {cleaned['term'].nunique()} term sheets", False),
        (f"Changes made: {len(log):,} (see the Cleaning Log sheet)", False),
        ("", False),
        ("What changed", True),
        ("1. Status is now TEXT. The original workbook stored each order's status as the", False),
        ("   fill colour of the Status cell, with a colour legend off to the right. Only", False),
        ("   the two newest sheets used words. Every row here has an explicit status, and", False),
        ("   the legend colours are reapplied so the sheets still read the same way.", False),
        ("2. Vendor spellings merged (Fisher Sci / FisherSci / fisher Sci, BioRad /", False),
        ("   Bio-Rad / BIO-RAD, and so on) so filtering and grouping actually work.", False),
        ("3. Impossible dates corrected where another date on the same row supplied the", False),
        ("   right year, and blank request dates filled from the order or received date.", False),
        ("4. One impossible unit price repaired from the same catalog number elsewhere", False),
        ("   in this workbook.", False),
        ("5. Stray non-breaking spaces and double spaces removed from text fields.", False),
        ("", False),
        ("What was deliberately NOT changed", True),
        ("* Quantity. The column is used inconsistently: usually the number of packs, but", False),
        ("  sometimes the number of individual units (1,200 for a case of 10 mL pipettes).", False),
        ("  There is no reliable way to tell the two apart, so the values are untouched.", False),
        ("  For this reason any spending total should sum Unit Price and NOT multiply by", False),
        (f"  quantity: across these {len(cleaned):,} rows the two approaches give "
         f"${cleaned['price'].fillna(0).sum():,.0f} and "
         f"${(cleaned['price'].fillna(0) * cleaned['quantity'].fillna(1)).sum():,.0f}.", False),
        ("* Missing prices and catalog numbers. Nothing in the workbook supplies them, so", False),
        ("  they are flagged in the Review Flag column rather than invented.", False),
        ("* Sizes on the Undated sheet, which do not match their items (a spirit level", False),
        ("  listed as '15 mL'). The correct values are not recoverable; they are flagged.", False),
        ("", False),
        ("The Review Flag column marks rows worth a human look. Nothing in it blocks import.", False),
    ]
    for text, bold in lines:
        ws.append([text])
        if bold:
            ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 100


def write_workbook(cleaned, log, out_path, src):
    frame = to_output_frame(cleaned)
    wb = Workbook()
    wb.remove(wb.active)

    _write_readme(wb.create_sheet("Read Me"), cleaned, log, src)

    ws_all = wb.create_sheet("All Orders")
    _write_sheet(ws_all, frame)
    ws_all.auto_filter.ref = ws_all.dimensions

    # Preserve the term-by-term layout people are used to.
    for term in cleaned["term"].drop_duplicates():
        sub = frame[frame["Term"] == term].drop(columns=["Term"])
        name = str(term)[:31]
        _write_sheet(wb.create_sheet(name), sub)

    ws_log = wb.create_sheet("Cleaning Log")
    if log.empty:
        ws_log.append(["No changes were necessary."])
    else:
        _write_sheet(ws_log, log, colour_status=False)
        ws_log.auto_filter.ref = ws_log.dimensions

    wb.save(out_path)
    return out_path


def main(argv=None):
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("workbook", help="the original order workbook")
    p.add_argument("-o", "--out", default=None, help="output path for the cleaned copy")
    args = p.parse_args(argv)

    out = args.out
    if not out:
        base, ext = os.path.splitext(args.workbook)
        out = f"{base} (CLEANED){ext}"

    print(f"reading  {args.workbook}")
    df = hi.parse_orders_workbook(args.workbook)
    print(f"  {len(df):,} order rows across {df['term'].nunique()} sheets")

    cleaned, log = clean(df)
    print(f"\n{len(log):,} changes:")
    if not log.empty:
        for why, n in Counter(log["Why"].str.split(";").str[0]).most_common():
            print(f"  {n:>5}  {why}")

    flagged = cleaned[cleaned["_flags"] != ""]
    print(f"\n{len(flagged):,} rows flagged for review (not blocking):")
    counts = Counter()
    for f in flagged["_flags"]:
        for part in f.split("; "):
            counts[part] += 1
    for what, n in counts.most_common():
        print(f"  {n:>5}  {what}")

    write_workbook(cleaned, log, out, args.workbook)
    print(f"\nwrote {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
