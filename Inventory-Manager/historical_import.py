"""Import the lab's historical Excel order workbook into the inventory database.

The workbook ("Sarmazdeh's Lab Orders.xlsx") is one sheet per term/fiscal year,
and it stores the real order status two different ways:

* On the older sheets the status is the FILL COLOUR of the Status cell (each
  sheet carries a colour legend off to the right), and the Status *text* column
  actually holds who received the item and when ("AB 28MAY2021").
* On the newest sheets (Fall 2025, Spring 2026) the Status text holds a real
  status word and the colour is unused.

So a recognised status word in the text wins; otherwise the fill colour decides;
otherwise the text is kept as a receipt note.

Everything imported is tagged with ``import_batch`` so a bad run can be undone in
one step, and with ``is_historical`` so the live dashboards can filter it out.
"""
import io
import re
from collections import Counter
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# --------------------------------------------------------------------------
# Workbook parsing
# --------------------------------------------------------------------------

SKIP_SHEETS = {"Instructions", "Pickup Duty", "Order List",
               # Meta sheets in a workbook produced by clean_workbook.py
               "Read Me", "Cleaning Log"}

# clean_workbook.py writes a combined sheet carrying its own Term column. When
# it is present it is the single source of truth — parsing the per-term sheets
# as well would double every row.
COMBINED_SHEET = "All Orders"

# Colour legend (fill of the Status cell -> meaning), mapped onto the status
# vocabulary admin_dashboard.py already uses.
COLOR_STATUS = {
    "rgb:FF00B050": "Received",
    "rgb:FFFFFF00": "Ordered",          # Ordered/Pending/Waiting for Shipment
    "rgb:FFFFC000": "Sent to Dr. MRS",
    "rgb:FFFF0000": "Need to order",
    "rgb:FFC00000": "Need to order",    # darker red, same intent
    "rgb:FFEEECE1": "Cancelled",        # Delayed/cancelled/misc.
    "rgb:FF0070C0": "Back order",
    "rgb:FF00B0F0": "Needs Fixing",
    "theme:7:0.0": "Do not order yet",  # WAIT
    "theme:5:0.0": "Lost",              # LOST
}

# Neutral greys used for the "Delayed/cancelled/misc." bucket. Only a status
# when the shading is on the Status cell alone — whole-row shading is banding.
GREY_FILLS = {"rgb:FFD8D8D8", "rgb:FFE7E6E6", "rgb:FFF2F2F2",
              "theme:0:-0.05", "theme:0:-0.1", "theme:0:-0.15"}
# Explicit white == "no colour applied".
NULL_FILLS = {"theme:0:0.0", "rgb:FFFFFFFF", "rgb:00000000"}

STATUS_VOCAB = {
    "received": "Received",
    "ordered": "Ordered",
    "shipped": "Shipped",
    "pending": "Pending",
    "waiting for shipment": "Waiting for Shipment",
    "sent to dr. mrs": "Sent to Dr. MRS",
    "sent to dr mrs": "Sent to Dr. MRS",
    "delayed": "Delayed",
    "back order": "Back order",
    "backorder": "Back order",
    "b/o": "Back order",
    "needs fixing": "Needs Fixing",
    "needs to be fixed": "Needs Fixing",
    "misc": "Misc.",
    "do not order yet": "Do not order yet",
    "wait": "Do not order yet",
    "cancelled": "Cancelled",
    "canceled": "Cancelled",
    "not needed": "Cancelled",
    "unavailable": "Cancelled",
    "discontinued": "Cancelled",
    "refunded": "Cancelled",
    "returned": "Cancelled",
    "lost": "Lost",
    "need to order": "Need to order",
}

ALIASES = {
    "item": "item", "size": "size", "quantity": "quantity",
    "cat no.": "catalog_number", "cat no": "catalog_number",
    "vendor": "seller", "unit price": "price", "initial": "initial",
    "date requested": "date_requested",
    "initial/date requested": "initial_date_requested",
    "initial/date received": "initial_date_received",
    "status": "status_text", "order date": "order_date",
    "order #": "order_number", "gr #": "gr_number",
    "expected delivery date": "expected_delivery",
    "tracking #": "tracking_number",
    "storage": "storage", "storage conditions": "storage",
    "comments": "comments", "link": "link",
    "received (initial/date)": "received_note", "stored": "stored",
    "no.": "row_no",
    # Extra columns produced by clean_workbook.py
    "term": "term_col", "received by": "received_by_col",
    "received date": "received_date_col", "source row": "source_row",
    "review flag": "review_flag",
}

DATE_FMTS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y",
             "%d-%b-%y", "%d-%b-%Y", "%b %d, %Y", "%d%b%Y", "%d%b%y")


def _fill_key(cell):
    f = cell.fill
    if f is None or f.patternType is None:
        return None
    c = f.fgColor
    try:
        if c.type == "rgb":
            return "rgb:" + str(c.rgb)
        if c.type == "theme":
            return f"theme:{c.theme}:{round(float(c.tint or 0), 3)}"
        if c.type == "indexed":
            return "idx:" + str(c.indexed)
    except Exception:
        pass
    return None


def _norm_header(v):
    if v is None:
        return None
    return ALIASES.get(re.sub(r"\s+", " ", str(v)).strip().lower().rstrip(":"))


def _find_header_row(ws):
    best, best_n = None, 0
    for r in range(1, min(ws.max_row, 12) + 1):
        n = sum(1 for c in range(1, ws.max_column + 1)
                if _norm_header(ws.cell(row=r, column=c).value))
        if n > best_n:
            best, best_n = r, n
    return best if best_n >= 4 else None


def _clean(v):
    if v is None or isinstance(v, datetime):
        return v
    if isinstance(v, str):
        v = v.replace("­", "").replace("�", "")
        return re.sub(r"\s+", " ", v).strip() or None
    return v


def _text(v):
    v = _clean(v)
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() or None


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return float(t) if t not in ("", "-", ".") else None
    except ValueError:
        return None


def _date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    t = str(v).strip()
    for fmt in DATE_FMTS:
        try:
            return datetime.strptime(t, fmt)
        except ValueError:
            pass
    for pat, fmts in ((r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
                       ("%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y")),
                      (r"(\d{1,2}-[A-Za-z]{3}-\d{2,4})", ("%d-%b-%y", "%d-%b-%Y"))):
        m = re.search(pat, t)
        if m:
            for fmt in fmts:
                try:
                    return datetime.strptime(m.group(1), fmt)
                except ValueError:
                    pass
    m = re.search(r"(\d{1,2}\s*[A-Za-z]{3}\s*\d{2,4})", t)   # 28MAY2021
    if m:
        cand = re.sub(r"\s+", "", m.group(1))
        for fmt in ("%d%b%Y", "%d%b%y"):
            try:
                return datetime.strptime(cand, fmt)
            except ValueError:
                pass
    return None


def _split_initial_date(v):
    """FY sheets pack 'AH 6/26/20' or 'AB 28MAY2021' into a single cell."""
    if v is None:
        return None, None
    if isinstance(v, datetime):
        return None, v
    t = str(v).strip()
    d = _date(t)
    initials = t
    for pat in (r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}",
                r"\d{1,2}-[A-Za-z]{3}-\d{2,4}",
                r"\d{1,2}\s*[A-Za-z]{3}\s*\d{2,4}"):
        initials = re.sub(pat, "", initials)
    return (initials.strip(" ,/-") or None), d


def _resolve_status(text_val, status_fill, item_fill):
    """Return (status, source, leftover_text). See the module docstring."""
    txt = _text(text_val)
    if txt:
        key = re.sub(r"\s+", " ", txt).strip().lower().rstrip(".!")
        hit = STATUS_VOCAB.get(key)
        if hit:
            return hit, "text", None
    if status_fill and status_fill not in NULL_FILLS:
        if status_fill in COLOR_STATUS:
            return COLOR_STATUS[status_fill], "color", txt
        if status_fill in GREY_FILLS and status_fill != item_fill:
            return "Cancelled", "color-grey", txt
    return None, "none", txt


def parse_orders_workbook(source, skip_sheets=SKIP_SHEETS):
    """Parse the order workbook into a DataFrame, one row per order line.

    ``source`` may be a path or a file-like object (e.g. a Streamlit upload).
    """
    if hasattr(source, "read"):
        source = io.BytesIO(source.read())
    wb = load_workbook(source, data_only=True)

    if COMBINED_SHEET in wb.sheetnames:
        sheets = [COMBINED_SHEET]
    else:
        sheets = [s for s in wb.sheetnames if s not in skip_sheets]

    records = []
    for sh in sheets:
        ws = wb[sh]
        hr = _find_header_row(ws)
        if hr is None:
            continue
        colmap = {}
        for c in range(1, ws.max_column + 1):
            key = _norm_header(ws.cell(row=hr, column=c).value)
            if key and key not in colmap:
                colmap[key] = c
        if "item" not in colmap:
            continue
        status_col, item_col = colmap.get("status_text"), colmap["item"]

        for r in range(hr + 1, ws.max_row + 1):
            def get(k):
                return _clean(ws.cell(row=r, column=colmap[k]).value) if k in colmap else None

            item = _text(get("item"))
            if not item or item.lower() in ("item", "no orders at this time"):
                continue

            status_fill = _fill_key(ws.cell(row=r, column=status_col)) if status_col else None
            item_fill = _fill_key(ws.cell(row=r, column=item_col))
            status, src, leftover = _resolve_status(get("status_text"), status_fill, item_fill)

            initial = _text(get("initial"))
            requested = _date(get("date_requested"))
            if initial is None and "initial_date_requested" in colmap:
                i2, d2 = _split_initial_date(get("initial_date_requested"))
                initial, requested = initial or i2, requested or d2

            recv_by, recv_on = (None, None)
            if "initial_date_received" in colmap:
                recv_by, recv_on = _split_initial_date(get("initial_date_received"))
            # A cleaned workbook splits these out into their own columns.
            recv_by = recv_by or _text(get("received_by_col"))
            recv_on = recv_on or _date(get("received_date_col"))
            for extra in (leftover, get("received_note")):
                if extra is not None and not recv_on:
                    r2, d3 = _split_initial_date(extra)
                    recv_by, recv_on = recv_by or r2, recv_on or d3

            records.append({
                "term": _text(get("term_col")) or sh,
                "excel_row": r,
                "item": item,
                "size": _text(get("size")),
                "quantity": _num(get("quantity")),
                "catalog_number": _text(get("catalog_number")),
                "seller": _text(get("seller")),
                "price": _num(get("price")),
                "initial": initial,
                "date_requested": requested,
                "order_date": _date(get("order_date")),
                "order_number": _text(get("order_number")),
                "gr_number": _text(get("gr_number")),
                "expected_delivery": _date(get("expected_delivery")),
                "tracking_number": _text(get("tracking_number")),
                "storage": _text(get("storage")),
                "comments": _text(get("comments")),
                "link": _text(get("link")),
                "received_by": recv_by,
                "received_date": recv_on,
                "stored": _text(get("stored")),
                "status": status,
                "status_source": src,
            })

    df = pd.DataFrame(records)
    if df.empty:
        return df
    # A row with no colour and no status word is almost always an old received
    # line whose colour was cleared; fall back to the receipt evidence.
    got_receipt = df["received_date"].notna() | df["order_number"].notna()
    df.loc[df["status"].isna() & got_receipt, "status"] = "Received"
    df["status"] = df["status"].fillna("Unknown")
    return df


# --------------------------------------------------------------------------
# Category inference (only used for the archived legacy inventory rows)
# --------------------------------------------------------------------------

# Evaluated top to bottom, so the specific patterns come before the generic
# ones ("...Master Mix" must land in Kits before the enzyme rule claims it).
CATEGORY_RULES = [
    ("Kits", r"\bkit\b|miniprep|midiprep|maxiprep|megaprep|mini prep|"
             r"purification system|clean-?up system|assembly master ?mix|"
             r"master ?mix|assay system|\bprep[s]?\b"),
    ("Cells", r"competent (e\.? ?coli|cells)|\bneb ?5|10-?beta|bl21|dh5|"
              r"\bcell line\b|\byeast strain\b|\bebY|\bhek ?293|\bcho\b cells"),
    ("Media", r"\bmedi(a|um)\b|\bbroth\b|\blb\b|\bagar\b|\bsoc\b|\byp[dn]\b|terrific|"
              r"\bfbs\b|fetal bovine|\bserum\b|\bdpbs\b|\bmem\b|\bdmem\b|expression medium|"
              r"\bsd-?caa\b|trypsin-?edta|cellbanker"),
    ("Protein", r"antibod|anti-|\bigg\b|serum albumin|\bbsa\b|protein\b|enzyme|"
                r"proteinase|\btrypsin\b|streptavidin|strep,|avidin|lysozyme|"
                r"recombinant|\bdnase\b|\brnase\b|\bnuclease\b|exonuclease|ligase|"
                r"polymerase|phosphatase|kinase\b|zymolyase|pngase|enterokinase|"
                r"\bhf\b$|\b(bam|eco|hind|nde|xho|nhe|bsr|sac|sal|kpn|not|spe|xba)[a-z]*i+"
                r"(-hf)?\b|collagen|fibronectin|\bconjugate\b"),
    ("Plasmid", r"plasmid|\bpet\b|\bpcr\b primer|\boligo|\bgblock|\bdna\b ladder|"
                r"\bprimer[s]?\b|gene fragment|\bgibson\b"),
    ("Buffer", r"buffer|\btris\b|\bhepes\b|\bpbs\b|\btbs\b|\btae\b|\btbe\b|\bmops\b"),
    ("Chemical", r"\bacid\b|sodium|potassium|calcium|magnesium|chloride|sulfate|"
                 r"phosphate|ethanol|methanol|isopropanol|acetone|glycerol|urea|"
                 r"guanidine|imidazole|\bedta\b|\bsds\b|\bdtt\b|\btemed\b|glucose|"
                 r"sucrose|glycine|arginine|glutamate|\bl-[a-z]|ampicillin|kanamycin|"
                 r"chloramphenicol|\biptg\b|\bx-?gal\b|bleach|reagent|acrylamide|"
                 r"\bsalt\b|powder|\bstain\b|\bdye\b|substrate|\bmtt\b|trypan|"
                 r"\bdmso\b|tween|triton|\bnaoh\b|\bhcl\b|detergent|ethidium"),
    ("Glassware", r"borosilicate|erlenmeyer|graduated cylinder|glass (bottle|beaker|flask|tube)"),
    ("Equipment", r"centrifuge|balance\b|heater|shaker|incubator|freezer|\bpump\b|"
                  r"\bmeter\b|stirrer|\brack\b|holder|clamp|apparatus|scale\b|lighter|"
                  r"divider|electrode|\bprobe\b|timer|forceps|scissors|\bcart\b|"
                  r"\bled\b|power supply|sonicator|vortex|thermometer|thrmomtr|\bhood\b|"
                  r"refrigerator|\bcounter\b|detector|monitor\b|\bfpl?c\b|\blamp\b|"
                  r"freezing container|\bbin\b|batteries"),
    ("Consumables", r"tip[s]?\b|tube[s]?\b|pipet|\bplate[s]?\b|flask|dish|bottle|glove|"
                    r"filter|column|resin\b|cuvette|vial|bag[s]?\b|wipe|towel|foil|"
                    r"parafilm|syringe|needle|membrane|weigh|boat|reservoir|cap[s]?\b|"
                    r"beaker|cylinder|serological|slide|marker|\bpen[s]?\b|dialysis|"
                    r"cassette|transwell|insert|label|toothpick|wrap|underpad|"
                    r"cryo|\bbox(es)?\b|tubing|swab|\bcotton\b|\bpaper\b|"
                    r"microplate|microfuge|stopper|sponge|\btape\b|chamber"),
]


def infer_category(name):
    n = (name or "").lower()
    for cat, pattern in CATEGORY_RULES:
        if re.search(pattern, n):
            return cat
    return "Other"


# --------------------------------------------------------------------------
# Planning the import
# --------------------------------------------------------------------------

TERMINAL_STATUSES = {"Received", "Cancelled", "Lost"}


def _norm_name(v):
    if not v or str(v).lower() in ("nan", "none"):
        return ""
    v = re.sub(r"[™®°–—]", " ", str(v).lower())
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", v)).strip()


def _norm_cat(v):
    if not v or str(v).lower() in ("nan", "none"):
        return ""
    return re.sub(r"[^a-z0-9]", "", str(v).lower())


# The workbook contains a handful of year typos ("2-Sep-35" for "2-Sep-25",
# "15-Jul-05" for "15-Jul-23") that Excel faithfully stored as 1935 and 1905.
# Dates outside this window are treated as unusable so a typo in one column
# cannot outrank a good date in another.
MIN_PLAUSIBLE_YEAR = 2000
MAX_PLAUSIBLE_YEAR = datetime.now().year + 2


def _plausible(v):
    if v is None or pd.isna(v):
        return None
    ts = pd.to_datetime(v, errors="coerce")
    if pd.isna(ts):
        return None
    return ts if MIN_PLAUSIBLE_YEAR <= ts.year <= MAX_PLAUSIBLE_YEAR else None


def _row_date(row):
    """Best available *plausible* date for an order line."""
    for k in ("date_requested", "order_date", "received_date", "expected_delivery"):
        ts = _plausible(row.get(k))
        if ts is not None:
            return ts
    # Everything is implausible — fall back to the raw request date rather than
    # dropping the row's date entirely.
    for k in ("date_requested", "order_date"):
        v = row.get(k)
        if v is not None and pd.notna(v):
            return pd.to_datetime(v, errors="coerce")
    return None


def plan_import(existing_requests, rows, date_window_days=45):
    """Decide which parsed rows are new.

    An order line counts as *already in the database* only when an existing
    purchase request matches on catalog number (or item name) **and** sits
    within ``date_window_days`` of it. Matching on catalog number alone is far
    too loose across six years — the lab re-buys the same products every term,
    so a bare catalog hit means "we still stock this", not "this order is
    already recorded".
    """
    idx = {}
    for _, r in existing_requests.iterrows():
        d = pd.to_datetime(r.get("request_date"), errors="coerce")
        for key in (("cat", _norm_cat(r.get("catalog_number"))),
                    ("name", _norm_name(r.get("item_name")))):
            if key[1]:
                idx.setdefault(key, []).append(d)

    dupes, fresh = [], []
    for _, row in rows.iterrows():
        rd = _row_date(row)
        hit = False
        for key in (("cat", _norm_cat(row.get("catalog_number"))),
                    ("name", _norm_name(row.get("item")))):
            if not key[1] or key not in idx:
                continue
            for d in idx[key]:
                if pd.isna(d) or rd is None:
                    continue
                if abs((d - rd).days) <= date_window_days:
                    hit = True
                    break
            if hit:
                break
        (dupes if hit else fresh).append(row)

    return (pd.DataFrame(fresh).reset_index(drop=True),
            pd.DataFrame(dupes).reset_index(drop=True))


def summarize(new_rows, assume_received=True):
    """Human-readable counts for the preview screen."""
    if new_rows.empty:
        return {"total": 0, "as_used": 0, "by_status": {}, "by_term": {},
                "reclassified": 0}
    received = int((new_rows["status"] == "Received").sum())
    return {
        "total": int(len(new_rows)),
        # Under the default assumption every imported row is treated as used.
        "as_used": int(len(new_rows)) if assume_received else received,
        "reclassified": int(len(new_rows)) - received if assume_received else 0,
        "by_status": dict(Counter(new_rows["status"])),
        "by_term": dict(Counter(new_rows["term"])),
    }


# A unit price above this is almost certainly a mistyped cell (a catalog number
# pasted into Unit Price, say) rather than a real purchase.
SUSPICIOUS_PRICE = 20000.0


def find_data_issues(rows):
    """Flag source rows worth eyeballing before or after an import.

    These are problems in the workbook itself, not import failures — the rows
    still import, they just deserve a human look.
    """
    if rows.empty:
        return pd.DataFrame()

    issues = []
    for _, r in rows.iterrows():
        why = []
        price = _val(r.get("price"))
        if price is not None and price > SUSPICIOUS_PRICE:
            why.append(f"unit price ${price:,.0f} looks mistyped")
        if price in (None, 0):
            why.append("no price recorded")
        if not _val(r.get("catalog_number")):
            why.append("no catalog number")
        if _val(r.get("date_requested")) is None and _val(r.get("order_date")) is None:
            why.append("no date")
        for field in ("date_requested", "order_date", "received_date"):
            raw = r.get(field)
            if raw is not None and pd.notna(raw) and _plausible(raw) is None:
                why.append(f"{field.replace('_', ' ')} reads "
                           f"{pd.to_datetime(raw, errors='coerce'):%Y-%m-%d} — likely a year typo")
        if why:
            issues.append({
                "term": r.get("term"), "excel_row": r.get("excel_row"),
                "item": r.get("item"), "price": price,
                "quantity": _val(r.get("quantity")),
                "issues": "; ".join(why),
            })
    return pd.DataFrame(issues)


# --------------------------------------------------------------------------
# Applying the import
# --------------------------------------------------------------------------

def _specs(row):
    parts = []
    if row.get("size"):
        parts.append(f"Size: {row['size']}")
    if row.get("comments"):
        parts.append(f"Comments: {row['comments']}")
    if row.get("gr_number"):
        parts.append(f"GR #: {row['gr_number']}")
    return " | ".join(parts)


def _val(v, default=None):
    """Normalise pandas/NumPy blanks to plain Python values."""
    if v is None:
        return default
    try:
        if pd.isna(v):
            return default
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass
    return v


def build_records(new_rows, batch_id, mark_received_as_used=True,
                  assume_received=True):
    """Turn parsed rows into (purchase_request, inventory) record dicts.

    Every line becomes a purchase request so the order history is complete.

    With ``assume_received`` (the default), anything being imported is recorded
    as **Received and fully used**: if it is old enough to be in the workbook
    but is not already in the live database, the lab has been and gone through
    it. Each row therefore also gets an inventory row that is already spent —
    quantity 0, depleted, archived — which is what "mark it as used" means in
    this schema. Such rows stay fully searchable (the purchase wizard surfaces
    archived items under "Archived / Legacy Items") without touching live stock
    counts or the low-stock alerts.

    The workbook's own status is preserved in ``specs`` whenever it was not
    "Received", so a cancelled or lost order is still identifiable as one.
    """
    requests, items = [], []
    for _, row in new_rows.iterrows():
        rd = _row_date(row)
        # No date anywhere is recorded as NULL rather than backfilled with
        # today, which would drop six-year-old orders into the live window.
        request_date = rd.to_pydatetime() if rd is not None else None
        original_status = _val(row.get("status"), "Unknown")
        status = "Received" if assume_received else original_status
        price = _val(row.get("price"), 0.0) or 0.0
        qty = _val(row.get("quantity"), 1.0) or 1.0
        specs = _specs(row)
        if assume_received and original_status not in ("Received", "Unknown"):
            specs = (specs + " | " if specs else "") + \
                    f"Workbook status: {original_status}"

        requests.append({
            "requester_name": _val(row.get("initial"), "") or "Legacy Import",
            "item_name": _val(row.get("item"), ""),
            "specs": specs,
            "catalog_number": _val(row.get("catalog_number"), "") or "",
            "seller": _val(row.get("seller"), "") or "",
            "link": _val(row.get("link"), "") or "",
            "price": float(price),
            "quantity": float(qty),
            "keep_on_ice": False,
            "status": status,
            "shipping_number": _val(row.get("tracking_number"), "") or "",
            "courier": "",
            "order_number": _val(row.get("order_number"), "") or "",
            "request_date": request_date,
            "status_updated_at": request_date,
            "source_sheet": _val(row.get("term"), "") or "",
            "is_historical": True,
            "import_batch": batch_id,
        })

        if mark_received_as_used and status == "Received":
            recv = _val(row.get("received_date"))
            depleted_at = (pd.to_datetime(recv).to_pydatetime()
                           if recv is not None else request_date)
            items.append({
                "name": _val(row.get("item"), ""),
                "category": infer_category(row.get("item")),
                "source_type": "Purchased",
                "quantity": 0.0,
                "unit": "units",
                "reorder_threshold": 0.0,
                "location": _val(row.get("storage"), "") or "",
                "owner": "Lab",
                "catalog_number": _val(row.get("catalog_number"), "") or "",
                "seller": _val(row.get("seller"), "") or "",
                "link": _val(row.get("link"), "") or "",
                "specs": specs,
                "price": float(price),
                "date_added": request_date,
                "is_depleted": True,
                "last_depleted": depleted_at,
                "archived": True,
                "source_sheet": _val(row.get("term"), "") or "",
                "is_historical": True,
                "import_batch": batch_id,
            })
    return requests, items


def apply_import(db, requests, items, progress=None):
    """Bulk-insert the planned records. Returns (n_requests, n_items)."""
    n_req = _bulk_insert(db, "purchase_requests", requests, progress, "orders")
    n_inv = _bulk_insert(db, "inventory", items, progress, "legacy stock")
    return n_req, n_inv


def _bulk_insert(db, table, records, progress=None, label=""):
    if not records:
        return 0
    df = pd.DataFrame(records)
    # to_sql in one shot is dramatically faster than row-by-row inserts, and it
    # goes through the same engine both in SQLite and Supabase mode.
    df.to_sql(table, db.conn, if_exists="append", index=False,
              chunksize=200, method="multi")
    db.commit()
    if progress:
        progress(f"Inserted {len(df)} {label} rows into {table}.")
    return len(df)


def undo_import(db, batch_id):
    """Delete everything a given import batch created."""
    removed = {}
    for table in ("inventory", "purchase_requests"):
        if db.is_postgres:
            db.execute(f"DELETE FROM {table} WHERE import_batch = :b", {"b": batch_id})
        else:
            db.execute(f"DELETE FROM {table} WHERE import_batch = ?", (batch_id,))
        removed[table] = getattr(db, "rowcount", None)
    db.commit()
    return removed


def list_batches(db):
    """Every import batch present in the database, newest first."""
    q = """
        SELECT import_batch,
               COUNT(*)          AS rows,
               MIN(request_date) AS first_date,
               MAX(request_date) AS last_date
        FROM purchase_requests
        WHERE import_batch IS NOT NULL AND import_batch != ''
        GROUP BY import_batch
        ORDER BY MAX(request_date) DESC
    """
    try:
        return db.get_query_df(q)
    except Exception:
        return pd.DataFrame()


# --------------------------------------------------------------------------
# Command line
# --------------------------------------------------------------------------

def _main(argv=None):
    import argparse

    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("workbook", help="path to the Excel order workbook")
    p.add_argument("--apply", action="store_true",
                   help="actually write to the database (default is a dry run)")
    p.add_argument("--batch", default=None, help="batch label (default: xlsx-<timestamp>)")
    p.add_argument("--window", type=int, default=45,
                   help="duplicate-detection window in days (default 45)")
    p.add_argument("--no-mark-used", action="store_true",
                   help="do not create archived legacy inventory rows")
    p.add_argument("--keep-workbook-status", action="store_true",
                   help="keep each row's original status instead of recording everything "
                        "as Received and used up")
    p.add_argument("--undo", metavar="BATCH", help="delete everything a batch created, then exit")
    p.add_argument("--db", default=None, help="SQLite path (ignored when Supabase secrets exist)")
    args = p.parse_args(argv)

    from db_manager import AdvancedLabInventory
    db = AdvancedLabInventory(db_name=args.db)
    target = "Supabase (cloud)" if db.is_postgres else f"SQLite ({db.db_path})"
    print(f"target database: {target}")

    if args.undo:
        undo_import(db, args.undo)
        print(f"removed batch {args.undo}")
        return 0

    parsed = parse_orders_workbook(args.workbook)
    print(f"parsed {len(parsed)} order rows from {parsed['term'].nunique()} sheets")

    existing = db.get_query_df(
        "SELECT request_id, item_name, catalog_number, request_date FROM purchase_requests")
    assume = not args.keep_workbook_status
    new_rows, dupes = plan_import(existing, parsed, date_window_days=args.window)
    stats = summarize(new_rows, assume_received=assume)
    print(f"already recorded (skipping): {len(dupes)}")
    print(f"to import                  : {stats['total']}")
    print(f"  of which marked as used  : {0 if args.no_mark_used else stats['as_used']}")
    if assume:
        print(f"  recorded as Received     : {stats['total']} "
              f"({stats['reclassified']} had another status in the workbook, kept in specs)")
    print("  workbook statuses:")
    for status, n in sorted(stats["by_status"].items(), key=lambda kv: -kv[1]):
        print(f"    {status:<20} {n}")

    issues = find_data_issues(new_rows)
    if not issues.empty:
        print(f"\n{len(issues)} rows have gaps/oddities in the workbook (they still import)")

    if not args.apply:
        print("\nDRY RUN — nothing written. Re-run with --apply to commit.")
        return 0

    batch = args.batch or f"xlsx-{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    reqs, items = build_records(new_rows, batch,
                                mark_received_as_used=not args.no_mark_used,
                                assume_received=assume)
    n_req, n_inv = apply_import(db, reqs, items)
    print(f"\nimported {n_req} order records and {n_inv} legacy inventory rows as batch '{batch}'")
    print(f"undo with:  python historical_import.py <workbook> --undo {batch}")
    return 0


if __name__ == "__main__":
    raise SystemExit(_main())
